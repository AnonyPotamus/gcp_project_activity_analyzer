import argparse
import datetime
import logging
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
import warnings
from urllib3.exceptions import InsecureRequestWarning

# Suppress all warnings
warnings.filterwarnings("ignore")

# Specifically suppress the oauth2client warning
warnings.filterwarnings("ignore", "file_cache is only supported with oauth2client<4.0.0")

# Suppress Google API client warnings
logging.getLogger('googleapiclient.discovery_cache').setLevel(logging.ERROR)

from google.cloud import resourcemanager_v3
from google.cloud import storage
from google.oauth2.credentials import Credentials
from google.auth.exceptions import DefaultCredentialsError
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import googleapiclient.discovery
from googleapiclient.errors import HttpError

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def parse_arguments():
    parser = argparse.ArgumentParser(description="Find last activity date for GCP projects in a folder.")
    parser.add_argument("folder_id", help="GCP Folder ID")
    parser.add_argument("--output", default="project_activity.xlsx", help="Output Excel file name")
    parser.add_argument("--max_workers", type=int, default=10, help="Maximum number of worker threads")
    return parser.parse_args()

def get_credentials():
    try:
        credentials = Credentials.from_authorized_user_file(
            os.path.expanduser("~/.config/gcloud/application_default_credentials.json")
        )
        return credentials
    except (DefaultCredentialsError, FileNotFoundError):
        logging.error("Unable to find local credentials. Please run 'gcloud auth application-default login'")
        raise

def get_projects_recursive(folder_id, credentials):
    folders_client = resourcemanager_v3.FoldersClient(credentials=credentials)
    projects_client = resourcemanager_v3.ProjectsClient(credentials=credentials)
    projects = []

    def traverse_folder(folder_id):
        folder_filter = f"parent.type:folder parent.id:{folder_id}"
        request = resourcemanager_v3.SearchProjectsRequest(query=folder_filter)
        for project in projects_client.search_projects(request=request):
            projects.append(project.project_id)

        subfolder_request = resourcemanager_v3.ListFoldersRequest(parent=f"folders/{folder_id}")
        for folder in folders_client.list_folders(request=subfolder_request):
            traverse_folder(folder.name.split('/')[-1])

    traverse_folder(folder_id)
    return projects

def check_compute_last_activity(project_id, credentials):
    try:
        compute = googleapiclient.discovery.build('compute', 'v1', credentials=credentials)
        
        # Check VM instances
        request = compute.instances().aggregatedList(project=project_id)
        response = request.execute()
        
        latest_activity = None
        
        for zone, instances in response.get('items', {}).items():
            if 'instances' in instances:
                for instance in instances['instances']:
                    # Parse creation timestamp
                    created = datetime.datetime.fromisoformat(instance['creationTimestamp'].replace('Z', '+00:00'))
                    if latest_activity is None or created > latest_activity:
                        latest_activity = created
                    
                    # Parse last start timestamp if available
                    if 'lastStartTimestamp' in instance:
                        last_start = datetime.datetime.fromisoformat(instance['lastStartTimestamp'].replace('Z', '+00:00'))
                        if latest_activity is None or last_start > latest_activity:
                            latest_activity = last_start
        
        return latest_activity, None
    except HttpError as e:
        if e.resp.status == 403 and 'accessNotConfigured' in str(e):
            return None, "API not enabled"
        logging.debug(f"Error checking compute activity for project {project_id}: {str(e)}")
        return None, str(e)
    except Exception as e:
        logging.debug(f"Error checking compute activity for project {project_id}: {str(e)}")
        return None, str(e)

def check_storage_last_activity(project_id, credentials):
    try:
        client = storage.Client(project=project_id, credentials=credentials)
        latest_activity = None
        
        # Get all buckets
        buckets = list(client.list_buckets())
        for bucket in buckets:
            # Check bucket creation time
            if latest_activity is None or bucket.time_created > latest_activity:
                latest_activity = bucket.time_created
                
            # Check bucket update time
            if bucket.updated and (latest_activity is None or bucket.updated > latest_activity):
                latest_activity = bucket.updated
                
            # Check last modified objects (limit to 10 most recent to avoid excessive API calls)
            blobs = list(client.list_blobs(bucket.name, max_results=10))
            for blob in blobs:
                if blob.updated and (latest_activity is None or blob.updated > latest_activity):
                    latest_activity = blob.updated
        
        return latest_activity, None
    except Exception as e:
        if 'accessNotConfigured' in str(e):
            return None, "API not enabled"
        logging.debug(f"Error checking storage activity for project {project_id}: {str(e)}")
        return None, str(e)

def check_api_usage_last_activity(project_id, credentials):
    try:
        # Use Service Usage API to check for recent API calls
        service = googleapiclient.discovery.build('serviceusage', 'v1', credentials=credentials)
        request = service.services().list(parent=f'projects/{project_id}', filter='state:ENABLED', pageSize=200)
        response = request.execute()
        
        if 'services' in response:
            # Just check if services are enabled, as a proxy for activity
            # Unfortunately, the API doesn't provide exact usage timestamps
            return datetime.datetime.now(datetime.timezone.utc), None
        
        return None, None
    except HttpError as e:
        if e.resp.status == 403 and 'accessNotConfigured' in str(e):
            return None, "API not enabled"
        logging.debug(f"Error checking API usage for project {project_id}: {str(e)}")
        return None, str(e)
    except Exception as e:
        logging.debug(f"Error checking API usage for project {project_id}: {str(e)}")
        return None, str(e)

def get_project_last_activity(project_id, credentials):
    try:
        project_info = {"project_id": project_id}
        
        # Initialize with None
        last_activity_date = None
        activity_source = "None"
        access_issues = []
        
        # Check different services and find the most recent activity
        compute_activity, compute_error = check_compute_last_activity(project_id, credentials)
        if compute_error:
            access_issues.append(f"Compute: {compute_error}")
        if compute_activity and (last_activity_date is None or compute_activity > last_activity_date):
            last_activity_date = compute_activity
            activity_source = "Compute Engine"
        
        storage_activity, storage_error = check_storage_last_activity(project_id, credentials)
        if storage_error:
            access_issues.append(f"Storage: {storage_error}")
        if storage_activity and (last_activity_date is None or storage_activity > last_activity_date):
            last_activity_date = storage_activity
            activity_source = "Cloud Storage"
        
        # If we don't have activity from specific services, check for general API usage
        if last_activity_date is None:
            api_activity, api_error = check_api_usage_last_activity(project_id, credentials)
            if api_error:
                access_issues.append(f"API: {api_error}")
            if api_activity:
                last_activity_date = api_activity
                activity_source = "API Usage"
        
        # Set the project info
        project_info["last_activity_date"] = last_activity_date
        project_info["activity_source"] = activity_source
        project_info["access_issues"] = "; ".join(access_issues) if access_issues else None
        
        return project_info
    except Exception as e:
        logging.error(f"Error getting last activity for project {project_id}: {str(e)}")
        return {
            "project_id": project_id, 
            "last_activity_date": None, 
            "activity_source": "Error",
            "access_issues": str(e)
        }

def write_to_excel(projects_activity, filename):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Project Activity"

        # Write headers
        headers = ['Project ID', 'Last Activity Date', 'Activity Source', 'Days Since Activity', 'Access Issues']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        # Define color for access issues
        access_issue_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow
        
        now = datetime.datetime.now(datetime.timezone.utc)
        
        # Write data
        for row, project in enumerate(projects_activity, start=2):
            ws.cell(row=row, column=1, value=project['project_id'])
            
            if project['last_activity_date']:
                activity_date = project['last_activity_date']
                ws.cell(row=row, column=2, value=activity_date.strftime('%Y-%m-%d %H:%M:%S'))
                
                # Calculate days since activity
                days_since = (now - activity_date).days
                ws.cell(row=row, column=4, value=days_since)
            else:
                ws.cell(row=row, column=2, value="No activity found")
                ws.cell(row=row, column=4, value="N/A")
                
            ws.cell(row=row, column=3, value=project['activity_source'])
            
            if 'access_issues' in project and project['access_issues']:
                ws.cell(row=row, column=5, value=project['access_issues'])
                # Highlight rows with access issues
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = access_issue_fill

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filename)
        logging.info(f"Results written to {filename}")
    except Exception as e:
        logging.error(f"Error writing to Excel: {str(e)}")

def main():
    args = parse_arguments()
    
    try:
        logging.info(f"Searching for projects in folder {args.folder_id}")
        
        credentials = get_credentials()
        projects = get_projects_recursive(args.folder_id, credentials)
        logging.info(f"Found {len(projects)} projects")
        
        # Get last activity for each project
        projects_activity = []
        with ThreadPoolExecutor(max_workers=args.max_workers) as executor:
            future_to_project = {executor.submit(get_project_last_activity, project_id, credentials): project_id 
                                 for project_id in projects}
            
            for future in as_completed(future_to_project):
                result = future.result()
                if result:
                    projects_activity.append(result)
                    logging.info(f"Processed project {result['project_id']}")
        
        # Sort projects by last activity date (most recent first)
        projects_activity.sort(key=lambda x: (x['last_activity_date'] is None, 
                                             x['last_activity_date'] if x['last_activity_date'] else datetime.datetime.min),
                              reverse=True)
        
        write_to_excel(projects_activity, args.output)
        
    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    main()